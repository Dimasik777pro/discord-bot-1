import discord
from discord import app_commands
from discord.ext import commands
import openpyxl
import os
from datetime import datetime

# === Настройки ===
TOKEN = os.getenv("DISCORD_TOKEN")  # на Replit токен хранится в Secrets
OWNER_ID = 123456789012345678  # <-- замени на свой Discord ID

# === Подготовка базы данных ===
db_file = "database.xlsx"

def init_db():
    if not os.path.exists(db_file):
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.title = "Users"
        sheet.append(["ID", "Name", "Дата"])
        wb.save(db_file)

def add_user(user: discord.User):
    wb = openpyxl.load_workbook(db_file)
    sheet = wb["Users"]
    # Проверка
    for row in sheet.iter_rows(values_only=True):
        if row[0] == str(user.id):
            wb.close()
            return
    # Добавление
    date_str = datetime.now().strftime("%d.%m.%Y %H:%M:%S")
    sheet.append([str(user.id), user.name, date_str])
    wb.save(db_file)
    wb.close()
    print(f"➕ Новый пользователь: {user.id} | {user.name} | {date_str}")

def get_users():
    wb = openpyxl.load_workbook(db_file)
    sheet = wb["Users"]
    data = []
    for row in sheet.iter_rows(values_only=True):
        if row[0] != "ID":  # пропустить заголовок
            data.append(row)
    wb.close()
    return data

def get_user_info(user: discord.User):
    wb = openpyxl.load_workbook(db_file)
    sheet = wb["Users"]
    for row in sheet.iter_rows(values_only=True):
        if row[0] == str(user.id):
            wb.close()
            return row  # (ID, Name, Дата)
    wb.close()
    return None

# === Бот ===
intents = discord.Intents.default()
bot = commands.Bot(command_prefix="!", intents=intents)

@bot.event
async def on_ready():
    print(f"✅ Бот {bot.user} запущен!")
    try:
        synced = await bot.tree.sync()  # глобальная синхронизация
        print(f"Синхронизировано {len(synced)} команд (глобально)")
    except Exception as e:
        print(f"Ошибка синхронизации: {e}")

# === Команды ===
@bot.tree.command(name="ping", description="Проверить пинг бота")
async def ping(interaction: discord.Interaction):
    add_user(interaction.user)
    await interaction.response.send_message(f"🏓 Понг! Задержка: {round(bot.latency*1000)}мс")

@bot.tree.command(name="info", description="Информация о пользователе")
async def info(interaction: discord.Interaction):
    add_user(interaction.user)
    user_info = get_user_info(interaction.user)
    if user_info:
        uid, name, date = user_info
        await interaction.response.send_message(
            f"👤 Твой ID: {uid}\nИмя: {name}\n📅 Добавлен в базу: {date}"
        )
    else:
        await interaction.response.send_message("❌ Ошибка: пользователь не найден в базе.")

@bot.tree.command(name="db", description="Показать базу данных (только владелец)")
async def db(interaction: discord.Interaction):
    if interaction.user.id != OWNER_ID:
        await interaction.response.send_message("❌ У тебя нет прав для этой команды!", ephemeral=True)
        return
    
    users = get_users()
    msg = "**📋 Список пользователей:**\n"
    for uid, name, date in users:
        msg += f"🔹 {uid} | {name} | {date}\n"

    # если сообщение слишком длинное
    if len(msg) > 1900:
        msg = "⚠ Слишком много пользователей, список не помещается в сообщение."

    await interaction.response.send_message(content=msg, file=discord.File(db_file))

# === Запуск ===
if __name__ == "__main__":
    init_db()
    from keep_alive import keep_alive
    keep_alive()
    bot.run(TOKEN)

